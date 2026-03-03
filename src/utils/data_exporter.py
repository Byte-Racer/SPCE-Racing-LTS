"""
Data Exporter — CSV and Excel export for lap and endurance results.

Exports telemetry arrays and aggregate statistics to structured files
that can be opened in Excel, MATLAB, or Python/pandas without any
simulator dependencies.

No regenerative braking is modelled in this simulator.

Author: Arceus, SPCE Racing
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

from src.solver.lap_results import EnduranceResult, LapResult

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Battery-power helper (shared with TelemetryPlotter logic)
# ---------------------------------------------------------------------------

def _battery_power_kw(lap: LapResult) -> np.ndarray:
    """Return per-segment battery power in kW (length N-1).

    Args:
        lap: Completed lap result.

    Returns:
        1-D ``float64`` array, length ``N-1``.
    """
    n = len(lap.distance)
    if n < 2 or len(lap.soc_profile) < 2:
        return np.zeros(max(n - 1, 0))

    soc_drop = -np.diff(lap.soc_profile)
    total_drop = lap.soc_profile[0] - lap.soc_profile[-1]

    if total_drop > 1e-9:
        energy_seg_wh = soc_drop / total_drop * lap.energy_consumed
    else:
        energy_seg_wh = np.zeros(n - 1)

    v_avg = (lap.speed_profile[:-1] + lap.speed_profile[1:]) / 2.0
    v_avg = np.maximum(v_avg, 0.1)
    dt_s = np.diff(lap.distance) / v_avg

    power_kw = energy_seg_wh * 3_600.0 / dt_s / 1_000.0
    return np.maximum(power_kw, 0.0)


class DataExporter:
    """Export simulation results to CSV and Excel files.

    Args:
        results_dir: Directory for output files (created if absent).

    Example::

        from src.utils.data_exporter import DataExporter
        exporter = DataExporter("results/run_01/")
        path = exporter.export_lap_csv(lap_result, "lap_telemetry.csv")
    """

    def __init__(self, results_dir: str = "results/") -> None:
        self.results_dir = Path(results_dir)
        self.results_dir.mkdir(parents=True, exist_ok=True)
        logger.info("DataExporter initialised → %s", self.results_dir)

    # ------------------------------------------------------------------ #
    #  CSV export                                                         #
    # ------------------------------------------------------------------ #

    def export_lap_csv(self, lap_result: LapResult, filename: str) -> str:
        """Export per-point lap telemetry to CSV.

        Columns: ``distance_m``, ``speed_kmh``, ``ax_g``,
        ``battery_power_kw``, ``soc_pct``, ``motor_temp_c``,
        ``limiting_factor``.

        Segment-length arrays (``ax_g``, ``battery_power_kw``,
        ``limiting_factor``) are right-padded with ``NaN`` / ``''``
        to align with point-length arrays.

        Args:
            lap_result: Completed lap telemetry container.
            filename:   Output filename (may be bare name or relative path).

        Returns:
            Absolute path string of the written file.
        """
        lap = lap_result
        n = len(lap.distance)

        # ── Pad segment arrays to length N ───────────────────────────
        if len(lap.ax_profile):
            ax_g = np.append(lap.ax_profile / 9.81, np.nan)[:n]
        else:
            ax_g = np.full(n, np.nan)

        power_kw_seg = _battery_power_kw(lap)
        if len(power_kw_seg):
            power_kw = np.append(power_kw_seg, np.nan)[:n]
        else:
            power_kw = np.full(n, np.nan)

        if lap.limiting_factor_profile:
            lf = (lap.limiting_factor_profile + [""])[:n]
        else:
            lf = [""] * n

        df = pd.DataFrame({
            "distance_m":        lap.distance,
            "speed_kmh":         lap.speed_profile * 3.6,
            "ax_g":              ax_g,
            "battery_power_kw":  power_kw,
            "soc_pct":           lap.soc_profile * 100.0,
            "motor_temp_c":      lap.motor_temp_profile,
            "limiting_factor":   lf,
        })

        out_path = self.results_dir / filename
        df.to_csv(out_path, index=False, float_format="%.4f")
        logger.info("Lap CSV written → %s  (%d rows)", out_path, len(df))
        return str(out_path)

    def export_endurance_csv(
            self, endurance_result: EnduranceResult, filename: str
    ) -> str:
        """Export per-lap endurance summary to CSV.

        One row per lap, plus summary rows appended at the bottom.

        Columns: ``lap_number``, ``lap_time_s``, ``energy_wh``,
        ``final_soc_pct``, ``final_motor_temp_c``, ``thermal_derating``,
        ``avg_speed_kmh``.

        Args:
            endurance_result: Completed endurance container.
            filename:         Output filename.

        Returns:
            Absolute path string of the written file.
        """
        rows = []
        for i, lap in enumerate(endurance_result.laps, start=1):
            rows.append({
                "lap_number":       i,
                "lap_time_s":       round(lap.lap_time, 4),
                "energy_wh":        round(lap.energy_consumed, 2),
                "final_soc_pct":    round(lap.final_soc * 100.0, 2),
                "final_motor_temp_c": round(lap.final_motor_temp, 2),
                "thermal_derating": lap.thermal_derating_occurred,
                "avg_speed_kmh":    round(lap.avg_speed * 3.6, 2),
            })

        df = pd.DataFrame(rows)

        # ── Summary rows ─────────────────────────────────────────────
        r = endurance_result
        summary_rows = [
            {
                "lap_number": "TOTAL",
                "lap_time_s": round(r.total_time, 4),
                "energy_wh": round(r.total_energy, 2),
                "final_soc_pct": round(r.final_soc * 100.0, 2),
                "final_motor_temp_c": "",
                "thermal_derating": "",
                "avg_speed_kmh": "",
            },
            {
                "lap_number": "AVG",
                "lap_time_s": round(r.avg_lap_time, 4),
                "energy_wh": round(r.total_energy / max(len(r.laps), 1), 2),
                "final_soc_pct": "",
                "final_motor_temp_c": "",
                "thermal_derating": "",
                "avg_speed_kmh": "",
            },
            {
                "lap_number": "BEST",
                "lap_time_s": round(r.best_lap_time, 4),
                "energy_wh": "",
                "final_soc_pct": "",
                "final_motor_temp_c": "",
                "thermal_derating": "",
                "avg_speed_kmh": "",
            },
            {
                "lap_number": "WORST",
                "lap_time_s": round(r.worst_lap_time, 4),
                "energy_wh": "",
                "final_soc_pct": "",
                "final_motor_temp_c": "",
                "thermal_derating": "",
                "avg_speed_kmh": "",
            },
        ]
        df_summary = pd.DataFrame(summary_rows)
        df_full = pd.concat([df, df_summary], ignore_index=True)

        out_path = self.results_dir / filename
        df_full.to_csv(out_path, index=False, float_format="%.4f")
        logger.info(
            "Endurance CSV written → %s  (%d laps + summary)",
            out_path, len(endurance_result.laps),
        )
        return str(out_path)

    # ------------------------------------------------------------------ #
    #  Excel report                                                       #
    # ------------------------------------------------------------------ #

    def export_excel_report(
            self,
            filename: str,
            lap_result: Optional[LapResult] = None,
            endurance_result: Optional[EnduranceResult] = None,
    ) -> str:
        """Export results to a multi-sheet Excel workbook.

        Sheets
        ------
        * **Lap Telemetry** — per-point data (if *lap_result* given)
        * **Lap Summary** — scalar KPIs for the lap
        * **Endurance Laps** — per-lap table with conditional formatting
        * **Vehicle Config** — placeholder sheet for config snapshot

        Requires ``openpyxl``.  Import is wrapped in ``try/except``
        so that the rest of the simulator works without it installed.

        Args:
            filename:         Output ``.xlsx`` filename.
            lap_result:       Optional lap result to include.
            endurance_result: Optional endurance result to include.

        Returns:
            Absolute path string of the written file.

        Raises:
            ImportError: If ``openpyxl`` is not installed.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError(
                "Install openpyxl: pip install openpyxl"
            ) from exc

        out_path = self.results_dir / filename

        red_fill    = PatternFill(start_color="FFCCCC",
                                  end_color="FFCCCC", fill_type="solid")
        orange_fill = PatternFill(start_color="FFE5B4",
                                  end_color="FFE5B4", fill_type="solid")
        header_font = Font(bold=True)

        def _auto_width(ws) -> None:
            """Set column widths to fit content."""
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        def _style_header_row(ws, row: int = 1) -> None:
            """Bold + freeze the header row."""
            for cell in ws[row]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = ws.cell(row=row + 1, column=1)

        with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:

            # ── Sheet 1: Lap Telemetry ─────────────────────────────────
            if lap_result is not None:
                n = len(lap_result.distance)

                if len(lap_result.ax_profile):
                    ax_g = np.append(lap_result.ax_profile / 9.81, np.nan)[:n]
                else:
                    ax_g = np.full(n, np.nan)

                pwr_seg = _battery_power_kw(lap_result)
                if len(pwr_seg):
                    pwr_col = np.append(pwr_seg, np.nan)[:n]
                else:
                    pwr_col = np.full(n, np.nan)

                if lap_result.limiting_factor_profile:
                    lf = (lap_result.limiting_factor_profile + [""])[:n]
                else:
                    lf = [""] * n

                df_tel = pd.DataFrame({
                    "distance_m":        lap_result.distance,
                    "speed_kmh":         lap_result.speed_profile * 3.6,
                    "ax_g":              ax_g,
                    "battery_power_kw":  pwr_col,
                    "soc_pct":           lap_result.soc_profile * 100.0,
                    "motor_temp_c":      lap_result.motor_temp_profile,
                    "limiting_factor":   lf,
                })
                df_tel.to_excel(writer, sheet_name="Lap Telemetry",
                                index=False, float_format="%.4f")
                ws_tel = writer.sheets["Lap Telemetry"]
                _style_header_row(ws_tel)
                _auto_width(ws_tel)

            # ── Sheet 2: Lap Summary ───────────────────────────────────
            if lap_result is not None:
                lap = lap_result
                summary_data = {
                    "Metric": [
                        "Lap Time (s)",
                        "Avg Speed (km/h)",
                        "Max Speed (km/h)",
                        "Min Speed (km/h)",
                        "Energy Consumed (Wh)",
                        "Final SOC (%)",
                        "Final Motor Temp (°C)",
                        "Thermal Derating",
                        "Energy Critical",
                    ],
                    "Value": [
                        round(lap.lap_time, 3),
                        round(lap.avg_speed * 3.6, 2),
                        round(lap.max_speed * 3.6, 2),
                        round(lap.min_speed * 3.6, 2),
                        round(lap.energy_consumed, 1),
                        round(lap.final_soc * 100.0, 1),
                        round(lap.final_motor_temp, 1),
                        str(lap.thermal_derating_occurred),
                        str(lap.energy_critical),
                    ],
                }
                df_sum = pd.DataFrame(summary_data)
                df_sum.to_excel(writer, sheet_name="Lap Summary",
                                index=False)
                ws_sum = writer.sheets["Lap Summary"]
                _style_header_row(ws_sum)
                _auto_width(ws_sum)

            # ── Sheet 3: Endurance Laps ────────────────────────────────
            if endurance_result is not None:
                rows = []
                for i, lap_e in enumerate(endurance_result.laps, start=1):
                    rows.append({
                        "lap_number":           i,
                        "lap_time_s":           round(lap_e.lap_time, 4),
                        "energy_wh":            round(lap_e.energy_consumed, 2),
                        "final_soc_pct":        round(lap_e.final_soc * 100.0, 2),
                        "final_motor_temp_c":   round(lap_e.final_motor_temp, 2),
                        "thermal_derating":     lap_e.thermal_derating_occurred,
                        "avg_speed_kmh":        round(lap_e.avg_speed * 3.6, 2),
                    })
                df_end = pd.DataFrame(rows)
                df_end.to_excel(writer, sheet_name="Endurance Laps",
                                index=False)
                ws_end = writer.sheets["Endurance Laps"]
                _style_header_row(ws_end)
                _auto_width(ws_end)

                # Conditional formatting — col indices (1-based in openpyxl)
                # Columns: lap_number=1, lap_time_s=2, energy_wh=3,
                #          final_soc_pct=4, final_motor_temp_c=5,
                #          thermal_derating=6, avg_speed_kmh=7
                for row_idx, lap_e in enumerate(endurance_result.laps, start=2):
                    temp_cell = ws_end.cell(row=row_idx, column=5)
                    soc_cell  = ws_end.cell(row=row_idx, column=4)
                    if isinstance(temp_cell.value, (int, float)) and temp_cell.value > 90:
                        for col in range(1, 8):
                            ws_end.cell(row=row_idx, column=col).fill = red_fill
                    elif isinstance(soc_cell.value, (int, float)) and soc_cell.value < 30:
                        for col in range(1, 8):
                            ws_end.cell(row=row_idx, column=col).fill = orange_fill

            # ── Sheet 4: Vehicle Config ────────────────────────────────
            df_cfg = pd.DataFrame({
                "Note": ["Vehicle configuration snapshot not available "
                         "in this export context.  Re-run with "
                         "SimulationReportGenerator.generate_*_report() "
                         "to include full YAML parameters."],
            })
            df_cfg.to_excel(writer, sheet_name="Vehicle Config", index=False)
            ws_cfg = writer.sheets["Vehicle Config"]
            _style_header_row(ws_cfg)
            _auto_width(ws_cfg)

        logger.info("Excel report written → %s", out_path)
        return str(out_path)
